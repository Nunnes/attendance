import openpyxl
import datetime
import databaseManager
import os 
from os import listdir
from os.path import isfile, join

data_path = None 

def init_attendance_excel_file(local_data_path):
	global data_path
	data_path = local_data_path

	filename = str(datetime.datetime.now().year) + ".xlsx"
	filePath = data_path + "/" + filename
	build_atttendance_file(filePath)

	print("Check Sheet Names:")
	print(get_sheet_names(filePath))

	return filename

def print_file_list():
	onlyfiles = [f for f in listdir(data_path) if isfile(join(data_path, f))]
	print(onlyfiles)

def get_sheet_names(filePath):
	wb = openpyxl.load_workbook(filePath)
	return wb.get_sheet_names()

def write_attendance(rfid_code, conn):
	wb = openpyxl.load_workbook(get_current_filepath())
	ws = wb[get_current_month_ws_id()]
	
	date =  datetime.date.today().strftime('%d - %A')
	now_time = datetime.datetime.now().strftime('%H:%M:%S')

	attendance = (rfid_code, date, now_time)
	databaseManager.insert_attendance(conn, attendance)	

	for row in ws.rows:
		row_number = str(row[0].row)

		if row[0].value == date and row[1].value == rfid_code:
			save_now_time(ws, row_number, now_time)
			wb.save(get_current_filepath())

			print(ws['C' + row_number].value)
			print(ws['D' + row_number].value)
			print(ws['E' + row_number].value)
			print(ws['F' + row_number].value)
			return

		elif row[0].value == None:
			ws.delete_rows(row[0].row, amount=1)
			wb.save(get_current_filepath())	

	data = [date, rfid_code, now_time]
	ws.append(data)
	wb.save(get_current_filepath())



def set_now_time_in_col_row(ws,col, row,now_time):
	if ws[col + row].value == None:
		ws[col + row] = now_time
		return True
	else: 
		return False

def save_now_time(ws, cell_row, now_time):
	if set_now_time_in_col_row(ws,'C', cell_row, now_time):
		return

	if set_now_time_in_col_row(ws,'D', cell_row, now_time):
		return

	if set_now_time_in_col_row(ws,'E', cell_row, now_time):
		return

	if set_now_time_in_col_row(ws,'F', cell_row, now_time):
		return
	
	print("Attendance table is full")
	
def build_atttendance_file(filePath):
	print("Create attendace excel file: " + filePath)

	if os.path.isfile(filePath) == False:
		wb = openpyxl.Workbook()
		months = [datetime.date(2000, m, 1).strftime('%m - %B') for m in range(1, 13)]
			
		for m in months:
			ws = wb.create_sheet(m)
			#ws['A1'] = 'DATE_ID_PK'
			ws['A1'] = 'Date'
			ws['B1'] = 'ID'

			ws['C1'] = 'In'
			ws['D1'] = 'Out'
			ws['E1'] = 'In'
			ws['F1'] = 'Out'
			
		wb.remove_sheet(wb.worksheets[0])
		wb.save(get_current_filepath())
	else:
		print ("File already exists: " + filePath)

def get_current_filepath():
	return data_path + "/" + str(datetime.datetime.now().year) + ".xlsx"

def get_current_month_ws_id():
	return datetime.date(datetime.datetime.now().year, datetime.datetime.now().month, 1).strftime('%m - %B')